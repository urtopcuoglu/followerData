import os
import requests
import openpyxl
import pandas as pd
from bs4 import BeautifulSoup

data = []

while True:
    username = input("Instagram kullanıcı adını girin ('q' girerek çıkış yapın): ")

    if username == 'q':
        break

    url = f"https://www.instagram.com/{username}/"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, "html.parser")

    # Takipçi sayısını bulma (HTML yapısına göre değişebilir)
    followers = soup.find("meta", property="og:description")["content"]
    followers = followers.split(" Followers")[0]

    # Kullanıcı adını ve tam adını bulma (HTML yapısına göre değişebilir)
    username = soup.find("meta", property="og:title")["content"]
    fullname = soup.find("title").text.split("•")[0].strip()

    data.append([username, fullname, followers])

    print("Kullanıcı adı:", username)
    print("Tam adı:", fullname)
    print("Takipçi sayısı:", followers)
    print()

# Excel dosyasının oluşturulacağı konumu belirleyin

try:
    workbook = openpyxl.load_workbook('C:/Users/Umutcan/veriler.xlsx')
    sheet = workbook['Sheet1']
    row = sheet.max_row + 1
    column = 1

    for item in data:
        username, fullname, followers = item
        cell = sheet.cell(row=row, column=column)
        cell.value = username
        cell = sheet.cell(row=row, column=column + 1)
        cell.value = fullname
        cell = sheet.cell(row=row, column=column + 2)
        cell.value = followers
        row += 1

    workbook.save('C:/Users/Umutcan/veriler.xlsx')

    # Onay mesajı
    print("Veriler Excel dosyasına başarıyla eklendi.")

except Exception as e:
    # Hata mesajı
    print("Veri aktarımı sırasında HATA!")
    print("Hata Açıklaması:", str(e))

print("Programdan Çıkılıyor...")